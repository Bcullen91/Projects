import xlsxwriter
from openpyxl import load_workbook

wb1 = load_workbook(filename = 'MoC_AllUsers.xlsx')
sheetAllUsers = wb1['AllUsers']
wb2 = load_workbook(filename = 'MoC_ActiveUsers.xlsx')
sheetActiveUsers = wb2['AllTimePosts']
wb3 = xlsxwriter.Workbook('NonActiveUsers.xlsx')
worksheet = wb3.add_worksheet('NonActives')
row = 0
col = 0

totalUsers = sheetAllUsers.max_row
totalPosters = sheetActiveUsers.max_row


users = []
posters = []


for i in range(1, totalPosters):
    posters.append(sheetActiveUsers['A' + str(i)].value)

for i in range(1, totalUsers):
    users.append(sheetAllUsers['A' + str(i)].value)

inactives = list(set(users) - set(posters))

print(users)
print(len(users))
print(posters)
print(len(posters))
print(inactives)
print(len(inactives))

for i in inactives:
    worksheet.write(row, col, i)
    row += 1


wb3.close()